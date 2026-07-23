import os
import time
import json
import csv
import argparse

import torch
from torch.utils.data import DataLoader
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import (
    DATASETS,
    TRAIN_SITES,
    TEST_SITES,
    MODEL_OUTPUT_DIR,
    BATCH_SIZE,
    EPOCHS,
    LEARNING_RATE,
    WEIGHT_DECAY,
    ENCODER_CHANNELS,
    DECODER_CHANNELS,
    BOTTLENECK_CHANNELS,
    OUTPUT_CHANNELS,
    DEVICE,
    NUM_WORKERS,
    AUGMENT_TRAIN
)
from model import UNet
from dataset import list_tiles, train_val_split, estimate_positive_ratio, LandslideDataset
from losses import BCEDiceLoss
from metrics import binary_metrics, raw_confusion_counts, metrics_from_confusion
from plotting import plot_training_curves, plot_confusion_matrix

MAX_POS_WEIGHT = 50.0


def format_value(value, fmt):
    if value is None:
        return "-"
    if fmt is None:
        return str(value)
    return format(value, fmt)


def print_table(title, rows, columns):
    """Prints a simple aligned ASCII table to the console.
    columns: list of (key, header, format_spec) tuples."""

    print(f"\n{title}")

    widths = []
    for key, header, fmt in columns:
        cell_values = [format_value(row.get(key), fmt) for row in rows]
        widths.append(max([len(header)] + [len(v) for v in cell_values]))

    def format_row(values):
        return " | ".join(value.ljust(width) for value, width in zip(values, widths))

    print(format_row([header for _, header, _ in columns]))
    print("-+-".join("-" * width for width in widths))

    for row in rows:
        values = [format_value(row.get(key), fmt) for key, _, fmt in columns]
        print(format_row(values))


def build_dataloaders(dataset_name):
    pairs = list_tiles(dataset_name, TRAIN_SITES)

    if not pairs:
        raise RuntimeError(f"No tiles found for dataset: {dataset_name}")

    train_pairs, val_pairs = train_val_split(pairs)

    train_dataset = LandslideDataset(train_pairs, augment=AUGMENT_TRAIN)
    val_dataset = LandslideDataset(val_pairs, augment=False)

    train_loader = DataLoader(
        train_dataset,
        batch_size=BATCH_SIZE,
        shuffle=True,
        num_workers=NUM_WORKERS,
        pin_memory=(DEVICE == "cuda"),
        drop_last=True
    )

    val_loader = DataLoader(
        val_dataset,
        batch_size=BATCH_SIZE,
        shuffle=False,
        num_workers=NUM_WORKERS,
        pin_memory=(DEVICE == "cuda")
    )

    print(f"  Tiles -> train: {len(train_pairs)}, val: {len(val_pairs)}")

    return train_loader, val_loader, train_pairs


def run_epoch(model, loader, criterion, optimizer=None, desc=""):
    is_training = optimizer is not None
    model.train() if is_training else model.eval()

    total_loss = 0.0
    total_metrics = {"iou": 0.0, "dice": 0.0, "precision": 0.0, "recall": 0.0, "accuracy": 0.0}
    num_batches = 0

    context = torch.enable_grad() if is_training else torch.no_grad()

    progress_bar = tqdm(loader, desc=desc, unit="batch", leave=False)

    with context:
        for images, masks in progress_bar:
            images = images.to(DEVICE)
            masks = masks.to(DEVICE)

            if is_training:
                optimizer.zero_grad()

            logits = model(images)
            loss = criterion(logits, masks)

            if is_training:
                loss.backward()
                optimizer.step()

            batch_metrics = binary_metrics(logits, masks)

            total_loss += loss.item()
            for key, value in batch_metrics.items():
                total_metrics[key] += value

            num_batches += 1

            progress_bar.set_postfix({
                "loss": f"{total_loss / num_batches:.4f}",
                "iou": f"{total_metrics['iou'] / num_batches:.4f}"
            })

    avg_loss = total_loss / num_batches
    avg_metrics = {key: value / num_batches for key, value in total_metrics.items()}

    return avg_loss, avg_metrics


def evaluate_confusion(model, loader):
    """Aggregates raw TP/FP/FN/TN pixel counts across an entire loader,
    used for the final confusion matrix plot."""

    model.eval()
    total_tp = total_fp = total_fn = total_tn = 0.0

    with torch.no_grad():
        for images, masks in tqdm(loader, desc="Evaluating confusion matrix", unit="batch", leave=False):
            images = images.to(DEVICE)
            masks = masks.to(DEVICE)

            logits = model(images)
            tp, fp, fn, tn = raw_confusion_counts(logits, masks)

            total_tp += tp
            total_fp += fp
            total_fn += fn
            total_tn += tn

    return total_tp, total_fp, total_fn, total_tn


def save_history_csv(history, path):
    if not history:
        return

    fieldnames = list(history[0].keys())

    with open(path, "w", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(history)


def save_history_xlsx(history, path, best_epoch=None):
    """Exports the full per-epoch history (including train/val loss and
    every tracked metric) to a formatted Excel workbook, with the best
    epoch's row highlighted."""

    if not history:
        return

    headers = list(history[0].keys())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Training History"

    sheet.append(headers)

    header_font = Font(name="Arial", bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    best_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for row in history:
        sheet.append([row[header] for header in headers])

    epoch_col_index = headers.index("epoch") + 1

    for row_cells in sheet.iter_rows(min_row=2):
        is_best_row = best_epoch is not None and row_cells[epoch_col_index - 1].value == best_epoch

        for cell in row_cells:
            cell.font = Font(name="Arial")

            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

            if is_best_row:
                cell.fill = best_fill

    sheet.freeze_panes = "A2"

    for col_index, header in enumerate(headers, start=1):
        cell_lengths = [len(header)]
        for row in history:
            value = row[header]
            cell_lengths.append(len(f"{value:.4f}") if isinstance(value, float) else len(str(value)))

        column_letter = sheet.cell(row=1, column=col_index).column_letter
        sheet.column_dimensions[column_letter].width = max(cell_lengths) + 2

    workbook.save(path)


def train_on_dataset(dataset_name, in_channels):
    print(f"\n=== Training on dataset: {dataset_name} ({in_channels} input channels) ===")

    train_loader, val_loader, train_pairs = build_dataloaders(dataset_name)

    positive_ratio = estimate_positive_ratio(train_pairs)

    if positive_ratio > 0:
        pos_weight = min((1.0 - positive_ratio) / positive_ratio, MAX_POS_WEIGHT)
    else:
        pos_weight = 1.0

    print(f"  Positive pixel ratio: {positive_ratio:.5f} -> pos_weight: {pos_weight:.2f}")

    model = UNet(
        in_channels=in_channels,
        out_channels=OUTPUT_CHANNELS,
        encoder_channels=ENCODER_CHANNELS,
        decoder_channels=DECODER_CHANNELS,
        bottleneck_channels=BOTTLENECK_CHANNELS
    ).to(DEVICE)

    criterion = BCEDiceLoss(pos_weight=pos_weight).to(DEVICE)
    optimizer = torch.optim.Adam(model.parameters(), lr=LEARNING_RATE, weight_decay=WEIGHT_DECAY)

    output_dir = os.path.join(MODEL_OUTPUT_DIR, dataset_name)
    os.makedirs(output_dir, exist_ok=True)

    best_val_iou = -1.0
    best_epoch = -1
    history = []

    for epoch in range(1, EPOCHS + 1):
        start_time = time.time()

        train_desc = f"[{dataset_name}] Epoch {epoch}/{EPOCHS} (train)"
        val_desc = f"[{dataset_name}] Epoch {epoch}/{EPOCHS} (val)"

        train_loss, train_metrics = run_epoch(model, train_loader, criterion, optimizer, desc=train_desc)
        val_loss, val_metrics = run_epoch(model, val_loader, criterion, desc=val_desc)

        elapsed = time.time() - start_time

        print(
            f"[{dataset_name}] Epoch {epoch}/{EPOCHS} ({elapsed:.1f}s) | "
            f"train_loss={train_loss:.4f} train_iou={train_metrics['iou']:.4f} "
            f"train_acc={train_metrics['accuracy']:.4f} | "
            f"val_loss={val_loss:.4f} val_iou={val_metrics['iou']:.4f} "
            f"val_acc={val_metrics['accuracy']:.4f}"
        )

        history.append({
            "epoch": epoch,
            "train_loss": train_loss,
            "val_loss": val_loss,
            **{f"train_{key}": value for key, value in train_metrics.items()},
            **{f"val_{key}": value for key, value in val_metrics.items()}
        })

        if val_metrics["iou"] > best_val_iou:
            best_val_iou = val_metrics["iou"]
            best_epoch = epoch
            torch.save(model.state_dict(), os.path.join(output_dir, "best_model.pth"))
            print(f"  -> new best val IoU: {best_val_iou:.4f}, checkpoint saved")

    torch.save(model.state_dict(), os.path.join(output_dir, "final_model.pth"))

    with open(os.path.join(output_dir, "history.json"), "w") as history_file:
        json.dump(history, history_file, indent=2)

    save_history_csv(history, os.path.join(output_dir, "history.csv"))
    save_history_xlsx(history, os.path.join(output_dir, "history.xlsx"), best_epoch=best_epoch)
    plot_training_curves(history, output_dir)

    # Confusion matrix (and final reported metrics) use the best
    # checkpoint, not the last epoch's weights
    model.load_state_dict(torch.load(os.path.join(output_dir, "best_model.pth"), map_location=DEVICE))
    tp, fp, fn, tn = evaluate_confusion(model, val_loader)
    plot_confusion_matrix(tp, fp, fn, tn, output_dir)

    final_metrics = metrics_from_confusion(tp, fp, fn, tn)

    with open(os.path.join(output_dir, "confusion_matrix.json"), "w") as cm_file:
        json.dump({"tp": tp, "fp": fp, "fn": fn, "tn": tn, **final_metrics}, cm_file, indent=2)

    print(f"  Saved: history.json/csv/xlsx, training_curves.png, confusion_matrix.png/json -> {output_dir}")

    summary_row = {
        "dataset": dataset_name,
        "best_epoch": best_epoch,
        "accuracy": final_metrics["accuracy"],
        "precision": final_metrics["precision"],
        "recall": final_metrics["recall"],
        "iou": final_metrics["iou"],
        "dice": final_metrics["dice"]
    }

    print_table(
        f"Final validation metrics ({dataset_name}, best checkpoint from epoch {best_epoch})",
        [summary_row],
        columns=[
            ("dataset", "Dataset", None),
            ("best_epoch", "Best Epoch", None),
            ("accuracy", "Accuracy", ".4f"),
            ("precision", "Precision", ".4f"),
            ("recall", "Recall", ".4f"),
            ("iou", "IoU", ".4f"),
            ("dice", "Dice", ".4f")
        ]
    )

    return summary_row


def parse_args():
    parser = argparse.ArgumentParser(
        description="Train the U-Net model on one or all dataset variants."
    )
    parser.add_argument(
        "--model",
        type=int,
        choices=[1, 2, 3, 4],
        default=None,
        help="Which dataset variant to train (1-4). If omitted, all datasets are trained."
    )
    return parser.parse_args()


def select_datasets(model_number):
    if model_number is None:
        return DATASETS

    dataset_names = list(DATASETS.keys())
    selected_name = dataset_names[model_number - 1]

    return {selected_name: DATASETS[selected_name]}


def main():
    args = parse_args()

    print("Training configuration loaded")
    print("Device:", DEVICE)
    print("Train sites:", TRAIN_SITES)
    print("Test sites (held out, not used here):", TEST_SITES)
    print("Epochs:", EPOCHS, "| Batch size:", BATCH_SIZE, "| Learning rate:", LEARNING_RATE, "| Weight decay:", WEIGHT_DECAY)

    datasets_to_run = select_datasets(args.model)
    print("Datasets to train:", list(datasets_to_run.keys()))

    all_results = []

    for dataset_name, in_channels in datasets_to_run.items():
        summary_row = train_on_dataset(dataset_name, in_channels)
        all_results.append(summary_row)

    if len(all_results) > 1:
        print_table(
            "Final Results Across All Dataset Variants (best checkpoint per variant)",
            all_results,
            columns=[
                ("dataset", "Dataset", None),
                ("best_epoch", "Best Epoch", None),
                ("accuracy", "Accuracy", ".4f"),
                ("precision", "Precision", ".4f"),
                ("recall", "Recall", ".4f"),
                ("iou", "IoU", ".4f"),
                ("dice", "Dice", ".4f")
            ]
        )

    print("\nFinished training dataset variant(s)")


if __name__ == "__main__":
    main()
